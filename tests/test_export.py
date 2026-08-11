import csv

import pytest
from openpyxl import load_workbook

from eqd2_calculator.export import (
    create_calculation_record,
    export_calculation_to_csv,
    export_calculations_to_csv,
    export_calculations_to_excel,
)


def test_create_calculation_record() -> None:
    record = create_calculation_record(
        component="Brachytherapy",
        dose_per_fraction=8.0,
        number_of_fractions=3,
        alpha_beta=3.0,
        total_physical_dose=24.0,
        bed=88.0,
        eqd2=52.8,
    )

    assert record == {
        "component": "Brachytherapy",
        "dose_per_fraction_gy": 8.0,
        "number_of_fractions": 3,
        "alpha_beta_gy": 3.0,
        "total_physical_dose_gy": 24.0,
        "bed_gy": 88.0,
        "eqd2_gy": 52.8,
    }


def test_create_ebrt_calculation_record() -> None:
    record = create_calculation_record(
        component="EBRT",
        dose_per_fraction=2.0,
        number_of_fractions=25,
        alpha_beta=10.0,
        total_physical_dose=50.0,
        bed=60.0,
        eqd2=50.0,
    )

    assert record["component"] == "EBRT"
    assert record["number_of_fractions"] == 25
    assert record["total_physical_dose_gy"] == pytest.approx(50.0)
    assert record["bed_gy"] == pytest.approx(60.0)
    assert record["eqd2_gy"] == pytest.approx(50.0)


def test_calculation_record_contains_expected_fields() -> None:
    record = create_calculation_record(
        component="EBRT",
        dose_per_fraction=1.8,
        number_of_fractions=25,
        alpha_beta=3.0,
        total_physical_dose=45.0,
        bed=72.0,
        eqd2=43.2,
    )

    assert set(record) == {
        "component",
        "dose_per_fraction_gy",
        "number_of_fractions",
        "alpha_beta_gy",
        "total_physical_dose_gy",
        "bed_gy",
        "eqd2_gy",
    }


def test_export_calculation_to_csv(tmp_path) -> None:
    record = create_calculation_record(
        component="Brachytherapy",
        dose_per_fraction=8.0,
        number_of_fractions=3,
        alpha_beta=3.0,
        total_physical_dose=24.0,
        bed=88.0,
        eqd2=52.8,
    )

    output_file = tmp_path / "calculation.csv"

    returned_path = export_calculation_to_csv(
        record=record,
        file_path=output_file,
    )

    assert returned_path == output_file
    assert output_file.exists()


def test_export_calculations_to_csv(tmp_path) -> None:
    ebrt_record = create_calculation_record(
        component="EBRT",
        dose_per_fraction=2.0,
        number_of_fractions=25,
        alpha_beta=3.0,
        total_physical_dose=50.0,
        bed=83.33,
        eqd2=50.0,
    )

    brachytherapy_record = create_calculation_record(
        component="Brachytherapy",
        dose_per_fraction=8.0,
        number_of_fractions=3,
        alpha_beta=3.0,
        total_physical_dose=24.0,
        bed=88.0,
        eqd2=52.8,
    )

    output_file = tmp_path / "combined_treatment.csv"

    returned_path = export_calculations_to_csv(
        records=[ebrt_record, brachytherapy_record],
        file_path=output_file,
    )

    assert returned_path == output_file
    assert output_file.exists()

    contents = output_file.read_text(encoding="utf-8")

    assert "EBRT" in contents
    assert "Brachytherapy" in contents


def test_export_calculations_to_csv_rejects_empty_records(
    tmp_path,
) -> None:
    output_file = tmp_path / "empty.csv"

    with pytest.raises(
        ValueError,
        match="At least one calculation record is required.",
    ):
        export_calculations_to_csv(
            records=[],
            file_path=output_file,
        )

    assert not output_file.exists()


def test_export_calculations_to_excel(tmp_path) -> None:
    ebrt_record = create_calculation_record(
        component="EBRT",
        dose_per_fraction=2.0,
        number_of_fractions=25,
        alpha_beta=3.0,
        total_physical_dose=50.0,
        bed=83.33,
        eqd2=50.0,
    )

    brachytherapy_record = create_calculation_record(
        component="Brachytherapy",
        dose_per_fraction=8.0,
        number_of_fractions=3,
        alpha_beta=3.0,
        total_physical_dose=24.0,
        bed=88.0,
        eqd2=52.8,
    )

    output_file = tmp_path / "combined_treatment.xlsx"

    returned_path = export_calculations_to_excel(
        records=[ebrt_record, brachytherapy_record],
        file_path=output_file,
    )

    assert returned_path == output_file
    assert output_file.exists()

    workbook = load_workbook(output_file)
    worksheet = workbook["Radiotherapy Results"]

    assert worksheet.max_row == 3
    assert worksheet.max_column == 7
    assert worksheet["A1"].value == "component"
    assert worksheet["A2"].value == "EBRT"
    assert worksheet["A3"].value == "Brachytherapy"
    assert worksheet["G2"].value == pytest.approx(50.0)
    assert worksheet["G3"].value == pytest.approx(52.8)

    workbook.close()


def test_export_calculations_to_excel_rejects_empty_records(
    tmp_path,
) -> None:
    output_file = tmp_path / "empty.xlsx"

    with pytest.raises(
        ValueError,
        match="At least one calculation record is required.",
    ):
        export_calculations_to_excel(
            records=[],
            file_path=output_file,
        )

    assert not output_file.exists()


def test_exported_csv_contains_correct_headers_and_values(
    tmp_path,
) -> None:
    record = create_calculation_record(
        component="Brachytherapy",
        dose_per_fraction=8.0,
        number_of_fractions=3,
        alpha_beta=3.0,
        total_physical_dose=24.0,
        bed=88.0,
        eqd2=52.8,
    )

    output_file = tmp_path / "verified_calculation.csv"

    export_calculation_to_csv(
        record=record,
        file_path=output_file,
    )

    with output_file.open(
        mode="r",
        newline="",
        encoding="utf-8",
    ) as csv_file:
        rows = list(csv.DictReader(csv_file))

    assert len(rows) == 1
    assert list(rows[0].keys()) == [
        "component",
        "dose_per_fraction_gy",
        "number_of_fractions",
        "alpha_beta_gy",
        "total_physical_dose_gy",
        "bed_gy",
        "eqd2_gy",
    ]
    assert rows[0]["component"] == "Brachytherapy"
    assert float(rows[0]["dose_per_fraction_gy"]) == pytest.approx(8.0)
    assert int(rows[0]["number_of_fractions"]) == 3
    assert float(rows[0]["alpha_beta_gy"]) == pytest.approx(3.0)
    assert float(
        rows[0]["total_physical_dose_gy"]
    ) == pytest.approx(24.0)
    assert float(rows[0]["bed_gy"]) == pytest.approx(88.0)
    assert float(rows[0]["eqd2_gy"]) == pytest.approx(52.8)


def test_exported_excel_contains_correct_headers_and_values(
    tmp_path,
) -> None:
    record = create_calculation_record(
        component="Brachytherapy",
        dose_per_fraction=8.0,
        number_of_fractions=3,
        alpha_beta=3.0,
        total_physical_dose=24.0,
        bed=88.0,
        eqd2=52.8,
    )

    output_file = tmp_path / "verified_calculation.xlsx"

    export_calculations_to_excel(
        records=[record],
        file_path=output_file,
    )

    workbook = load_workbook(output_file)
    worksheet = workbook["Radiotherapy Results"]

    headers = [
        worksheet.cell(row=1, column=column).value
        for column in range(1, 8)
    ]
    values = [
        worksheet.cell(row=2, column=column).value
        for column in range(1, 8)
    ]

    assert headers == [
        "component",
        "dose_per_fraction_gy",
        "number_of_fractions",
        "alpha_beta_gy",
        "total_physical_dose_gy",
        "bed_gy",
        "eqd2_gy",
    ]
    assert values[0] == "Brachytherapy"
    assert values[1] == pytest.approx(8.0)
    assert values[2] == 3
    assert values[3] == pytest.approx(3.0)
    assert values[4] == pytest.approx(24.0)
    assert values[5] == pytest.approx(88.0)
    assert values[6] == pytest.approx(52.8)

    workbook.close()